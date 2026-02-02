"""
Django REST Views para AI Reports
"""

import asyncio
import json
from django.shortcuts import get_object_or_404
from django.http import StreamingHttpResponse
from django.utils import timezone
from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from rest_framework.parsers import JSONParser

from .models import ChatSession, ChatMessage, GeneratedReport, AIAgentConfig
from .serializers import (
    ChatSessionSerializer, ChatMessageSerializer, GeneratedReportSerializer,
    AIAgentConfigSerializer, AIReportRequestSerializer, AIReportResponseSerializer,
    AIReportStreamSerializer
)
from .agent import process_ai_request, ProcessingStage
from datetime import datetime

# Import RBAC utilities
from users.rbac_utils import user_has_permission, log_audit


class ChatSessionViewSet(viewsets.ModelViewSet):
    """
    ViewSet para gerenciar sessões de chat
    """
    serializer_class = ChatSessionSerializer
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        """Retornar apenas sessões do usuário autenticado"""
        return ChatSession.objects.filter(user=self.request.user)
    
    def perform_create(self, serializer):
        """Criar sessão para o usuário autenticado"""
        serializer.save(user=self.request.user)
    
    @action(detail=True, methods=['get'], url_path='messages')
    def get_messages(self, request, pk=None):
        """
        GET /api/ai-reports/chat-sessions/{id}/messages/
        Retorna todas as mensagens de uma sessão
        """
        session = self.get_object()
        messages = session.messages.all()
        serializer = ChatMessageSerializer(messages, many=True)
        return Response(serializer.data)
    
    @action(detail=True, methods=['post'], url_path='archive')
    def archive_session(self, request, pk=None):
        """
        POST /api/ai-reports/chat-sessions/{id}/archive/
        Arquiva uma sessão de chat
        """
        session = self.get_object()
        session.is_archived = True
        session.save()
        return Response({'status': 'sessão arquivada'})
    
    @action(detail=False, methods=['delete'], url_path='clear-all')
    def clear_all(self, request):
        """
        DELETE /api/ai-reports/chat-sessions/clear-all/
        Deleta todas as sessões do usuário
        """
        count, _ = ChatSession.objects.filter(user=request.user).delete()
        return Response({'deleted': count})


class ChatMessageViewSet(viewsets.ModelViewSet):
    """
    ViewSet para gerenciar mensagens de chat
    """
    serializer_class = ChatMessageSerializer
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        """Retornar apenas mensagens de sessões do usuário"""
        return ChatMessage.objects.filter(session__user=self.request.user)
    
    @action(detail=False, methods=['post'], url_path='send')
    def send_message(self, request):
        """
        POST /api/ai-reports/messages/send/
        Envia uma mensagem e processa com IA
        
        Body:
        {
            "message": "Analyze inventory by country",
            "session_id": 1
        }
        
        Requer permissões:
        - create_ai_reports: para criar novos relatórios
        - use_ai_agents: para usar agentes IA
        """
        # Verificar permissões RBAC
        if not user_has_permission(request.user, 'create_ai_reports'):
            log_audit(
                request.user,
                'permission_denied',
                'ChatMessage',
                description='Tentativa de criar relatório sem permissão'
            )
            return Response(
                {'error': 'You do not have permission to create AI reports'},
                status=status.HTTP_403_FORBIDDEN
            )
        
        if not user_has_permission(request.user, 'use_ai_agents'):
            log_audit(
                request.user,
                'permission_denied',
                'ChatMessage',
                description='Tentativa de usar agentes sem permissão'
            )
            return Response(
                {'error': 'You do not have permission to use AI agents'},
                status=status.HTTP_403_FORBIDDEN
            )
        
        serializer = AIReportRequestSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        
        message = serializer.validated_data['message']
        session_id = serializer.validated_data.get('session_id')
        agent_id = serializer.validated_data.get('agent_id')
        
        # Criar ou buscar sessão
        if session_id:
            session = get_object_or_404(ChatSession, id=session_id, user=request.user)
        else:
            session = ChatSession.objects.create(user=request.user)
        
        # Buscar agente configurado
        if agent_id:
            agent = get_object_or_404(AIAgentConfig, id=agent_id, is_active=True)
        else:
            # Usar agente padrão (primeiro ativo)
            agent = AIAgentConfig.objects.filter(is_active=True).first()
            if not agent:
                return Response(
                    {'error': 'Nenhum agente ativo configurado'},
                    status=status.HTTP_400_BAD_REQUEST
                )
        
        # Log de auditoria - criação de sessão/mensagem
        log_audit(
            request.user,
            'create',
            'ChatMessage',
            object_id=str(session.id),
            description=f'Criou mensagem com agente {agent.name}: {message[:100]}'
        )
        
        # Salvar mensagem do usuário
        user_message = ChatMessage.objects.create(
            session=session,
            message_type='user',
            content=message
        )
        
        # Processar com IA (usar asyncio.run para executar coroutine)
        try:
            state = asyncio.run(process_ai_request(
                user_request=message,
                user_id=str(request.user.id),
                session_id=str(session.id),
                agent_config=agent  # Passar a configuração do agente
            ))
            
            # Salvar resposta da IA com informação do agente
            ai_message = ChatMessage.objects.create(
                session=session,
                message_type='ai',
                content=state['report_title'],
                status='complete',
                agent=agent  # Salvar qual agente gerou a resposta
            )
            
            # Salvar relatório gerado
            if state.get('report_data'):
                GeneratedReport.objects.create(
                    session=session,
                    title=state['report_title'],
                    description=f"Gerado em resposta a: {message[:100]}",
                    report_data=state['report_data'],
                    insights=state.get('insights', [])
                )
            
            # Atualizar título da sessão se necessário
            if not session.title:
                session.title = message[:50]
                session.save()
            
            response_data = {
                'session_id': session.id,
                'user_message_id': user_message.id,
                'ai_message_id': ai_message.id,
                'report_title': state['report_title'],
                'report_data': state.get('report_data', {}),
                'insights': state.get('insights', []),
                'recommendations': state.get('recommendations', []),
                'stage_progress': state.get('stage_progress', []),
                'processing_times': state.get('processing_times', {})
            }
            
            return Response(response_data, status=status.HTTP_201_CREATED)
        
        except Exception as e:
            # Salvar mensagem de erro
            error_message = ChatMessage.objects.create(
                session=session,
                message_type='ai',
                content=f"Erro ao processar requisição: {str(e)}",
                status='complete'
            )
            
            return Response(
                {'error': str(e), 'message_id': error_message.id},
                status=status.HTTP_400_BAD_REQUEST
            )


class GeneratedReportViewSet(viewsets.ModelViewSet):
    """
    ViewSet para gerenciar relatórios gerados
    """
    serializer_class = GeneratedReportSerializer
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        """Retornar apenas relatórios de sessões do usuário"""
        return GeneratedReport.objects.filter(session__user=self.request.user)
    
    @action(detail=True, methods=['post'], url_path='export/pdf')
    def export_pdf(self, request, pk=None):
        """
        POST /api/ai-reports/reports/{id}/export/pdf/
        Exporta relatório em PDF
        """
        report = self.get_object()
        
        try:
            # Importar reportlab dinamicamente se disponível
            from reportlab.lib.pagesizes import letter, A4
            from reportlab.lib import colors
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from io import BytesIO
            
            # Criar PDF em memória
            pdf_buffer = BytesIO()
            doc = SimpleDocTemplate(pdf_buffer, pagesize=A4)
            elements = []
            
            # Título
            styles = getSampleStyleSheet()
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=24,
                textColor=colors.HexColor('#10b981'),
                spaceAfter=12
            )
            
            title = Paragraph(report.title, title_style)
            elements.append(title)
            
            # Descrição
            desc = Paragraph(report.description or "", styles['Normal'])
            elements.append(desc)
            elements.append(Spacer(1, 12))
            
            # KPIs
            kpis = report.report_data.get('kpis', {})
            if kpis:
                kpi_title = Paragraph("<b>KPIs Principais</b>", styles['Heading2'])
                elements.append(kpi_title)
                
                kpi_data = [[k, str(v)] for k, v in kpis.items()]
                kpi_table = Table(kpi_data)
                kpi_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#10b981')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 12),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 1, colors.grey)
                ]))
                elements.append(kpi_table)
                elements.append(Spacer(1, 12))
            
            # Insights
            insights = report.insights or []
            if insights:
                insights_title = Paragraph("<b>Insights Principais</b>", styles['Heading2'])
                elements.append(insights_title)
                
                for insight in insights:
                    insight_para = Paragraph(f"• {insight}", styles['Normal'])
                    elements.append(insight_para)
                elements.append(Spacer(1, 12))
            
            # Data de geração
            timestamp = Paragraph(
                f"<i>Gerado em {report.created_at.strftime('%d/%m/%Y às %H:%M')}</i>",
                styles['Normal']
            )
            elements.append(timestamp)
            
            # Construir PDF
            doc.build(elements)
            
            # Retornar PDF
            pdf_buffer.seek(0)
            response = StreamingHttpResponse(pdf_buffer.read(), content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="{report.title}.pdf"'
            
            # Registrar export
            report.exported_formats.append('pdf')
            report.save()
            
            return response
        
        except ImportError:
            return Response(
                {'error': 'reportlab não está instalado. Instale com: pip install reportlab'},
                status=status.HTTP_400_BAD_REQUEST
            )
    
    @action(detail=True, methods=['post'], url_path='export/excel')
    def export_excel(self, request, pk=None):
        """
        POST /api/ai-reports/reports/{id}/export/excel/
        Exporta relatório em Excel
        """
        report = self.get_object()
        
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            from io import BytesIO
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Report"
            
            # Título
            ws['A1'] = report.title
            ws['A1'].font = Font(size=16, bold=True, color="10B981")
            ws.merge_cells('A1:D1')
            
            # KPIs
            row = 3
            kpis = report.report_data.get('kpis', {})
            ws[f'A{row}'] = "KPIs"
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
            
            for k, v in kpis.items():
                ws[f'A{row}'] = k
                ws[f'B{row}'] = str(v)
                row += 1
            
            # Insights
            row += 1
            ws[f'A{row}'] = "Insights"
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
            
            for insight in report.insights or []:
                ws[f'A{row}'] = insight
                ws.row_dimensions[row].height = 30
                ws[f'A{row}'].alignment = Alignment(wrap_text=True)
                row += 1
            
            # Salvar em memória
            excel_buffer = BytesIO()
            wb.save(excel_buffer)
            excel_buffer.seek(0)
            
            response = StreamingHttpResponse(excel_buffer.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename="{report.title}.xlsx"'
            
            # Registrar export
            report.exported_formats.append('excel')
            report.save()
            
            return response
        
        except ImportError:
            return Response(
                {'error': 'openpyxl não está instalado. Instale com: pip install openpyxl'},
                status=status.HTTP_400_BAD_REQUEST
            )
    
    @action(detail=True, methods=['post'], url_path='export/json')
    def export_json(self, request, pk=None):
        """
        POST /api/ai-reports/reports/{id}/export/json/
        Exporta relatório em JSON
        """
        report = self.get_object()
        
        export_data = {
            'title': report.title,
            'description': report.description,
            'created_at': report.created_at.isoformat(),
            'report_data': report.report_data,
            'insights': report.insights
        }
        
        response = Response(export_data)
        response['Content-Disposition'] = f'attachment; filename="{report.title}.json"'
        
        # Registrar export
        report.exported_formats.append('json')
        report.save()
        
        return response


class AIAgentConfigViewSet(viewsets.ModelViewSet):
    """
    ViewSet para gerenciar configurações do agente IA (admin only)
    """
    queryset = AIAgentConfig.objects.all()
    serializer_class = AIAgentConfigSerializer
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        """Apenas usuários staff podem ver todas as configurações"""
        if self.request.user.is_staff:
            return AIAgentConfig.objects.all()
        return AIAgentConfig.objects.filter(is_active=True)
    
    @action(detail=False, methods=['get'], url_path='active-agents')
    def active_agents(self, request):
        """
        GET /api/ai-reports/agent-config/active-agents/
        Lista todos os agentes ativos disponíveis para o usuário
        """
        agents = AIAgentConfig.objects.filter(is_active=True)
        serializer = self.get_serializer(agents, many=True)
        return Response(serializer.data)

